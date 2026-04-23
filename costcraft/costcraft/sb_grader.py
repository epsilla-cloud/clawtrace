"""SpreadsheetBench deterministic grader.

Modes:
  - Cell-Level: compare value at answer_position cell(s) between agent output
    and gold answer for each test case.
  - Sheet-Level: cell-by-cell compare over each sheet in the gold answer.

Score: fraction of test cases passed (agent produced xlsx is applied per test
case — we just compare the agent's SINGLE output against test case 1's answer
for simplicity, noting this deviates slightly from the official multi-test-case
protocol. Documented as a limitation.).
"""
from __future__ import annotations
import re
from dataclasses import dataclass, asdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


@dataclass
class SBGradeItem:
    check: str               # description of the check
    passed: bool
    expected: str | None
    got: str | None


@dataclass
class SBGradeResult:
    task_id: str
    passed: bool
    normalized: float        # 1.0 or 0.0 (Cell-Level) or fraction (Sheet-Level)
    items: list[SBGradeItem]
    failure_reason: str
    cells_checked: int
    cells_matched: int


def _find_agent_xlsx(deliverable_dir: Path) -> Path | None:
    if not deliverable_dir.exists():
        return None
    # Prefer files whose name contains "output" or matches {id}
    xlsx = sorted(deliverable_dir.rglob("*.xlsx"))
    if not xlsx:
        return None
    # Prefer one with "output" in name
    for p in xlsx:
        if "output" in p.name.lower() or "answer" in p.name.lower():
            return p
    return xlsx[0]


def _values_equal(a, b, tol: float = 1e-4) -> bool:
    if a == b:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) < tol
    except (TypeError, ValueError):
        pass
    return str(a).strip() == str(b).strip()


def _parse_answer_position(pos: str, default_sheet: str) -> tuple[str, str]:
    """Return (sheet, cell_or_range). Accepts 'Sheet!A1', "'Sheet 1'!A1:B5", 'A1:B5', or just a column like "'Sheet1'!F"."""
    if not pos:
        return default_sheet, ""
    pos = pos.strip()
    m = re.match(r"^'([^']+)'!(.+)$", pos)
    if m:
        return m.group(1), m.group(2)
    m = re.match(r"^([A-Za-z0-9_]+)!(.+)$", pos)
    if m:
        return m.group(1), m.group(2)
    return default_sheet, pos


def _iterate_range(ws, range_str: str):
    """Yield (coord, value) for cells in a range like 'A1', 'A1:B5', or 'B' (whole column)."""
    if re.match(r"^[A-Z]+$", range_str):
        # whole column, but limit to min(max_row, 200)
        col = range_str
        max_r = min(ws.max_row or 1, 200)
        for r in range(1, max_r + 1):
            cell = ws[f"{col}{r}"]
            yield cell.coordinate, cell.value
        return
    if re.match(r"^[A-Z]+[0-9]+$", range_str):
        cell = ws[range_str]
        yield cell.coordinate, cell.value
        return
    if ":" in range_str:
        try:
            for row in ws[range_str]:
                for cell in row:
                    yield cell.coordinate, cell.value
            return
        except Exception:
            pass
    # Fallback — treat as single cell
    try:
        yield range_str, ws[range_str].value
    except Exception:
        pass


def grade(task, deliverable_dir: Path) -> SBGradeResult:
    from .spreadsheetbench import SBTask
    assert isinstance(task, SBTask)

    agent_xlsx = _find_agent_xlsx(deliverable_dir)
    if agent_xlsx is None:
        return SBGradeResult(
            task_id=task.task_id, passed=False, normalized=0.0,
            items=[], failure_reason="no xlsx deliverable found",
            cells_checked=0, cells_matched=0,
        )

    if not task.test_cases:
        return SBGradeResult(
            task_id=task.task_id, passed=False, normalized=0.0,
            items=[], failure_reason="no test cases in task",
            cells_checked=0, cells_matched=0,
        )

    ans_xlsx = task.test_cases[0]["answer"]
    try:
        wb_g = load_workbook(agent_xlsx, data_only=True)
        wb_g_raw = load_workbook(agent_xlsx, data_only=False)
        wb_e = load_workbook(ans_xlsx, data_only=True)
    except Exception as e:
        return SBGradeResult(
            task_id=task.task_id, passed=False, normalized=0.0,
            items=[], failure_reason=f"xlsx load error: {e}",
            cells_checked=0, cells_matched=0,
        )

    default_sheet = wb_e.sheetnames[0]
    sheet, range_str = _parse_answer_position(task.answer_position, default_sheet)

    if task.instruction_type.lower().startswith("cell"):
        # Cell-level: compare at answer_position
        items = []
        checked = 0
        matched = 0
        if sheet not in wb_g.sheetnames or sheet not in wb_e.sheetnames:
            return SBGradeResult(
                task_id=task.task_id, passed=False, normalized=0.0,
                items=[], failure_reason=f"sheet missing: {sheet}",
                cells_checked=0, cells_matched=0,
            )
        ws_g = wb_g[sheet]
        ws_g_raw = wb_g_raw[sheet]
        for coord, exp_val in _iterate_range(wb_e[sheet], range_str):
            try:
                got_val = ws_g[coord].value
            except Exception:
                got_val = None
            # If computed value is None but there's a formula, use the raw form
            if got_val is None:
                try:
                    raw = ws_g_raw[coord].value
                    if isinstance(raw, str) and raw.startswith("="):
                        got_val = raw  # formula text — may still match if exp_val is formula
                except Exception:
                    pass
            checked += 1
            eq = _values_equal(got_val, exp_val)
            if eq:
                matched += 1
            items.append(SBGradeItem(
                check=f"{sheet}!{coord}", passed=eq,
                expected=str(exp_val) if exp_val is not None else None,
                got=str(got_val) if got_val is not None else None,
            ))
        passed = checked > 0 and matched == checked
        return SBGradeResult(
            task_id=task.task_id, passed=passed,
            normalized=(matched / checked) if checked > 0 else 0.0,
            items=items, failure_reason="" if passed else "cell mismatches",
            cells_checked=checked, cells_matched=matched,
        )
    else:
        # Sheet-level: compare full sheets
        items = []
        checked = 0
        matched = 0
        for sn in wb_e.sheetnames:
            if sn not in wb_g.sheetnames:
                items.append(SBGradeItem(check=f"sheet {sn}", passed=False, expected="exists", got="missing"))
                continue
            ws_e = wb_e[sn]
            ws_g = wb_g[sn]
            ws_gr = wb_g_raw[sn]
            for row in ws_e.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    checked += 1
                    got = ws_g[c.coordinate].value
                    if got is None:
                        raw = ws_gr[c.coordinate].value
                        if isinstance(raw, str) and raw.startswith("="):
                            got = raw
                    if _values_equal(got, c.value):
                        matched += 1
                    elif len(items) < 20:
                        items.append(SBGradeItem(
                            check=f"{sn}!{c.coordinate}",
                            passed=False,
                            expected=str(c.value),
                            got=str(got) if got is not None else None,
                        ))
        passed = checked > 0 and matched == checked
        return SBGradeResult(
            task_id=task.task_id,
            passed=passed,
            normalized=(matched / checked) if checked > 0 else 0.0,
            items=items,
            failure_reason="" if passed else f"{checked - matched}/{checked} cells mismatch",
            cells_checked=checked, cells_matched=matched,
        )


def grade_as_dict(task, deliverable_dir: Path) -> dict:
    r = grade(task, deliverable_dir)
    return {
        "task_id": r.task_id,
        "passed": r.passed,
        "normalized": r.normalized,
        "cells_checked": r.cells_checked,
        "cells_matched": r.cells_matched,
        "failure_reason": r.failure_reason,
        "items": [asdict(it) for it in r.items[:30]],
    }
