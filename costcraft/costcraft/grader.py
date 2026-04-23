"""GDPval rubric grader — single batched LLM-judge call per deliverable.

Uses claude-agent-sdk → Claude Code subscription (no API key). One batched call
per deliverable scores all rubric items at once, keeping rate-limit pressure
minimal. Falls back to per-item calls if JSON parsing fails.
"""
from __future__ import annotations
import asyncio
import json
import re
from dataclasses import dataclass
from pathlib import Path

from .claude import aoneshot


DELIVERABLE_SUMMARY_CHAR_LIMIT = 18000


@dataclass
class GradeItem:
    rubric_item_id: str
    criterion: str
    score: int        # points earned (0 or rubric's `score` value)
    max_score: int    # points possible
    passed: bool
    judge_rationale: str


@dataclass
class GradeResult:
    task_id: str
    total_points: int
    max_points: int
    normalized: float
    items: list[GradeItem]
    deliverable_summary_head: str
    judge_input_tokens: int
    judge_output_tokens: int
    judge_cost_usd: float


# ------------- deliverable summarization -------------

def summarize_xlsx(path: Path, max_rows: int = 60) -> str:
    """Render each cell as its cached computed value OR its formula text.

    Agents often save xlsx with formulas (`=H3-G3`); openpyxl in data_only mode
    shows None because Excel never cached the value. We fall back to the raw
    formula so the judge can evaluate it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return f"[openpyxl missing] {path.name}"
    try:
        wb_val = load_workbook(path, data_only=True)
        wb_raw = load_workbook(path, data_only=False)
    except Exception as e:
        return f"[xlsx load error: {e}] {path.name}"
    parts = [f"## XLSX: {path.name}"]
    for sheet in wb_val.sheetnames:
        ws_v = wb_val[sheet]
        ws_r = wb_raw[sheet]
        parts.append(f"### Sheet: {sheet}")
        for i, (row_v, row_r) in enumerate(zip(ws_v.iter_rows(values_only=True), ws_r.iter_rows(values_only=True))):
            if i >= max_rows:
                parts.append("... (truncated)")
                break
            cells = []
            any_non_empty = False
            for v, r in zip(row_v, row_r):
                if v is None and isinstance(r, str) and r.startswith("="):
                    cells.append(r)  # formula text
                    any_non_empty = True
                elif v is None and r is None:
                    cells.append("")
                else:
                    cells.append(str(v if v is not None else r))
                    any_non_empty = True
            if any_non_empty:
                parts.append("| " + " | ".join(cells) + " |")
    return "\n".join(parts)


def summarize_pdf(path: Path, max_chars: int = 8000) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return f"[pypdf missing] {path.name}"
    try:
        reader = PdfReader(str(path))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception as e:
        return f"[pdf load error: {e}] {path.name}"
    return f"## PDF: {path.name}\n\n{text[:max_chars]}"


def summarize_docx(path: Path, max_chars: int = 8000) -> str:
    try:
        import docx
    except ImportError:
        return f"[python-docx missing] {path.name}"
    try:
        doc = docx.Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        return f"[docx load error: {e}] {path.name}"
    return f"## DOCX: {path.name}\n\n{text[:max_chars]}"


def summarize_deliverable(deliverable_dir: Path) -> str:
    parts = []
    for f in sorted(deliverable_dir.rglob("*")):
        if not f.is_file() or f.name.startswith("."):
            continue
        sfx = f.suffix.lower()
        if sfx == ".xlsx":
            parts.append(summarize_xlsx(f))
        elif sfx == ".pdf":
            parts.append(summarize_pdf(f))
        elif sfx == ".docx":
            parts.append(summarize_docx(f))
        else:
            parts.append(f"## FILE: {f.name}  ({f.stat().st_size} bytes)")
    if not parts:
        return "[NO DELIVERABLE FILES FOUND]"
    joined = "\n\n".join(parts)
    return joined[:DELIVERABLE_SUMMARY_CHAR_LIMIT]


# ------------- batched LLM-judge -------------

BATCH_JUDGE_SYSTEM = (
    "You are a strict grader evaluating a deliverable against a list of rubric criteria.\n"
    "For EACH criterion, decide whether the deliverable satisfies it.\n"
    "Be literal: match values within ±0.01 if the criterion asks for numeric equality.\n"
    "Accept minor formatting differences (e.g. '$2.00' vs '2' vs '2.0') unless the criterion "
    "explicitly requires a formatting style.\n"
    "If the deliverable lacks the relevant item, return passed=false.\n\n"
    "Output: a JSON array where each entry matches the input order:\n"
    '  [{"id": "<rubric_item_id>", "passed": true|false, "rationale": "<=15 word reason"}, ...]\n'
    "Return ONLY the JSON array — no explanation, no code fences, no preface."
)


def _build_batch_user(deliverable_summary: str, rubric_json: list[dict]) -> str:
    criteria_lines = []
    for r in rubric_json:
        rid = str(r.get("rubric_item_id") or "")
        crit = str(r.get("criterion") or "").replace("\n", " ")
        criteria_lines.append(f"- id: {rid}\n  criterion: {crit}")
    criteria_block = "\n".join(criteria_lines)
    return (
        f"## DELIVERABLE\n\n{deliverable_summary}\n\n"
        f"## CRITERIA (score each)\n\n{criteria_block}\n\n"
        "Return a JSON array with one entry per criterion, in the same order."
    )


def _parse_batch_response(text: str, rubric_json: list[dict]) -> dict[str, tuple[bool, str]]:
    """Return {rubric_item_id: (passed, rationale)}. Missing items default to (False, 'no-judgment')."""
    t = text.strip()
    t = re.sub(r"^```(?:json)?\s*", "", t)
    t = re.sub(r"\s*```$", "", t)
    # If trailing junk, try to locate the JSON array
    if not t.startswith("["):
        m = re.search(r"\[[\s\S]*\]", t)
        if m:
            t = m.group(0)
    result: dict[str, tuple[bool, str]] = {}
    try:
        arr = json.loads(t)
        if isinstance(arr, list):
            for entry in arr:
                if not isinstance(entry, dict):
                    continue
                rid = str(entry.get("id") or "")
                passed = bool(entry.get("passed", False))
                rationale = str(entry.get("rationale", ""))[:200]
                if rid:
                    result[rid] = (passed, rationale)
    except Exception:
        pass
    return result


async def _judge_single(rid: str, criterion: str, deliverable_summary: str) -> tuple[bool, str]:
    """Per-item fallback when batched response is malformed for a specific id."""
    user = (
        f"## DELIVERABLE\n\n{deliverable_summary}\n\n"
        f"## CRITERION\n\n{criterion}\n\n"
        'Reply ONLY with JSON: {"passed": true|false, "rationale": "<=15 word reason"}.'
    )
    try:
        resp = await aoneshot(user=user, system=BATCH_JUDGE_SYSTEM)
        t = resp.text.strip()
        t = re.sub(r"^```(?:json)?\s*", "", t)
        t = re.sub(r"\s*```$", "", t)
        data = json.loads(t)
        return bool(data.get("passed", False)), str(data.get("rationale", ""))[:200]
    except Exception as e:
        return False, f"[fallback_error] {e}"


async def agrade_deliverable(
    *,
    task_id: str,
    rubric_json: list[dict],
    deliverable_dir: Path,
) -> GradeResult:
    summary = summarize_deliverable(deliverable_dir)

    # Fast-path: empty / missing deliverable → everything fails
    if summary.startswith("[NO DELIVERABLE FILES FOUND]"):
        return _build_all_fail(task_id, rubric_json, summary,
                               reason="no deliverable files produced")

    # Batched judge call (with retry on SDK failures)
    user = _build_batch_user(summary, rubric_json)
    try:
        resp = await aoneshot(user=user, system=BATCH_JUDGE_SYSTEM)
    except Exception as e:
        # Retry once with a shorter summary (in case size was the issue)
        try:
            user_short = _build_batch_user(summary[:8000], rubric_json)
            resp = await aoneshot(user=user_short, system=BATCH_JUDGE_SYSTEM)
        except Exception as e2:
            return _build_all_fail(task_id, rubric_json, summary,
                                   reason=f"judge_sdk_error: {e2}")
    parsed = _parse_batch_response(resp.text, rubric_json)

    # Fill in missing items via per-item fallback (bounded concurrency)
    missing_ids = [
        str(r.get("rubric_item_id") or "") for r in rubric_json
        if str(r.get("rubric_item_id") or "") not in parsed
    ]
    fallback_cost_extra = 0
    if missing_ids:
        sem = asyncio.Semaphore(4)

        async def _guarded(rid, crit):
            async with sem:
                return rid, await _judge_single(rid, crit, summary)

        coro = []
        for r in rubric_json:
            rid = str(r.get("rubric_item_id") or "")
            if rid in missing_ids:
                coro.append(_guarded(rid, str(r.get("criterion") or "")))
        fallbacks = await asyncio.gather(*coro)
        for rid, pair in fallbacks:
            parsed[rid] = pair

    items: list[GradeItem] = []
    total = 0
    max_total = 0
    for r in rubric_json:
        rid = str(r.get("rubric_item_id") or "")
        pts = int(r.get("score") or 1)
        crit = str(r.get("criterion") or "")
        passed, rationale = parsed.get(rid, (False, "[missing]"))
        earned = pts if passed else 0
        total += earned
        max_total += pts
        items.append(GradeItem(
            rubric_item_id=rid,
            criterion=crit,
            score=earned,
            max_score=pts,
            passed=passed,
            judge_rationale=rationale,
        ))
    norm = (total / max_total) if max_total > 0 else 0.0

    return GradeResult(
        task_id=task_id,
        total_points=total,
        max_points=max_total,
        normalized=norm,
        items=items,
        deliverable_summary_head=summary[:2000],
        judge_input_tokens=resp.input_tokens,
        judge_output_tokens=resp.output_tokens,
        judge_cost_usd=resp.total_cost_usd,
    )


def grade_deliverable(
    *, task_id: str, rubric_json: list[dict], deliverable_dir: Path
) -> GradeResult:
    return asyncio.run(agrade_deliverable(
        task_id=task_id, rubric_json=rubric_json, deliverable_dir=deliverable_dir
    ))


async def agrade_deliverable_n_times(
    *,
    task_id: str,
    rubric_json: list[dict],
    deliverable_dir: Path,
    n_rounds: int = 3,
) -> tuple[GradeResult, list[GradeResult]]:
    """Grade `n_rounds` times, return a majority-vote aggregate + the per-round list.

    Majority vote per rubric item: item passes if ≥⌈n/2⌉ rounds passed it.
    """
    rounds: list[GradeResult] = []
    for _ in range(n_rounds):
        r = await agrade_deliverable(
            task_id=task_id,
            rubric_json=rubric_json,
            deliverable_dir=deliverable_dir,
        )
        rounds.append(r)

    # Majority vote per rubric_item_id
    threshold = (n_rounds // 2) + 1
    item_votes: dict[str, int] = {}
    item_max: dict[str, int] = {}
    item_crit: dict[str, str] = {}
    item_rationales: dict[str, list[str]] = {}
    for rr in rounds:
        for it in rr.items:
            item_votes[it.rubric_item_id] = item_votes.get(it.rubric_item_id, 0) + (1 if it.passed else 0)
            item_max[it.rubric_item_id] = it.max_score
            item_crit[it.rubric_item_id] = it.criterion
            item_rationales.setdefault(it.rubric_item_id, []).append(it.judge_rationale)

    items: list[GradeItem] = []
    total = 0
    max_total = 0
    for rid, votes in item_votes.items():
        passed = votes >= threshold
        pts = item_max[rid]
        earned = pts if passed else 0
        total += earned
        max_total += pts
        items.append(GradeItem(
            rubric_item_id=rid,
            criterion=item_crit[rid],
            score=earned,
            max_score=pts,
            passed=passed,
            judge_rationale=f"majority {votes}/{n_rounds}: " + " | ".join(item_rationales[rid])[:300],
        ))

    agg = GradeResult(
        task_id=task_id,
        total_points=total,
        max_points=max_total,
        normalized=(total / max_total) if max_total > 0 else 0.0,
        items=items,
        deliverable_summary_head=rounds[0].deliverable_summary_head if rounds else "",
        judge_input_tokens=sum(r.judge_input_tokens for r in rounds),
        judge_output_tokens=sum(r.judge_output_tokens for r in rounds),
        judge_cost_usd=sum(r.judge_cost_usd for r in rounds),
    )
    return agg, rounds


def grade_deliverable_n_times(
    *, task_id: str, rubric_json: list[dict], deliverable_dir: Path, n_rounds: int = 3
) -> tuple[GradeResult, list[GradeResult]]:
    return asyncio.run(agrade_deliverable_n_times(
        task_id=task_id, rubric_json=rubric_json,
        deliverable_dir=deliverable_dir, n_rounds=n_rounds,
    ))


def _build_all_fail(task_id: str, rubric_json: list[dict],
                     summary: str, reason: str) -> GradeResult:
    """Record a 0-score grading result for a missing/unprocessable deliverable."""
    items: list[GradeItem] = []
    max_total = 0
    for r in rubric_json:
        pts = int(r.get("score") or 1)
        max_total += pts
        items.append(GradeItem(
            rubric_item_id=str(r.get("rubric_item_id") or len(items)),
            criterion=str(r.get("criterion") or ""),
            score=0,
            max_score=pts,
            passed=False,
            judge_rationale=reason[:200],
        ))
    return GradeResult(
        task_id=task_id,
        total_points=0,
        max_points=max_total,
        normalized=0.0,
        items=items,
        deliverable_summary_head=summary[:500],
        judge_input_tokens=0,
        judge_output_tokens=0,
        judge_cost_usd=0.0,
    )
